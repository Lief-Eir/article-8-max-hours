"""
Notes!
I would like to add a way to automatically download the dependencies, as explaining this is awful
I also would like to make it easier for user to load in the file to be parsed
This may include path cleaning.  Logic took like a few hours pathing and un-dummying is going to take forever
"""
# dependencies
from pandas import read_csv
from openpyxl import Workbook


class MaxHour:
    """ converts csv files to dataframe using pandas, detects violations and then outputs them to a spreadsheet
        using openpyxl"""
    def __init__(self):
        self.filein = None
        self.fileout = None
        self.df = None
        self.book = None
        self.sheet = None
        self.line_count = -1
        self.rowmax = -1
        self.xlrow = 2
        self.total = 0
        self.total12 = 0
        self.total60 = 0
        self.total602 = 0
        self.trutotal = 0

    def run(self):
        """ master method for running methods in proper sequence. """
        self.get_file()
        self.clean_filename()
        self.build_workbook()
        self.get_rowmax()
        self.main_loop()
        self.save()

    def get_file(self):
        """ ask for input.  It is on the user to know.  Make a README """
        self.filein = "docs/" + str(input('enter input file name: '))
        self.fileout = "docs/" + str(input('enter desired output file name: '))

    def clean_filename(self):
        """Input/Output Cleaning"""
        if not self.fileout.endswith('.xlsx'):
            self.fileout += '.xlsx'
        if not self.filein.endswith('.csv'):
            self.filein += '.csv'
        # Initialize Pandas DataTable, skipping the header
        self.df = read_csv(self.filein, skiprows=1)

    def build_workbook(self):
        """Initialize output with headers """
        self.book = Workbook()
        self.sheet = self.book.active
        self.sheet.cell(row=1, column=1).value = 'EID'
        self.sheet.cell(row=1, column=2).value = 'NAME'
        self.sheet.cell(row=1, column=3).value = 'TOTAL HOURS'
        self.sheet.cell(row=1, column=4).value = 'GRIEVED HOURS'

    # initialize counter variables in __init__

    def get_rowmax(self):
        """Find the number of rows to iterate over """
        for _ in self.df.iterrows():
            self.rowmax += 1

    def main_loop(self):
        """ All the logic """
        while self.line_count < self.rowmax - 4:
            self.line_count += 1
            self.total = 0
            self.total12 = 0
            self.check_craft()
            self.read()
            self.check_60()
            self.calculate()
            self.output()

    def check_craft(self):
        """ Checking craft """
        while self.df.at[self.line_count, 'D/A'] != 134:
            self.line_count += 1

    def read(self):
        """ We are verifying that the row is a full hour row for the same employee, then checking 12 hour violations """
        while self.df.at[self.line_count, 'Employee_ID'] == self.df.at[self.line_count + 1, 'Employee_ID'] and \
                self.df.at[self.line_count, 'Hours'] == 'WK':
            self.total += self.df.at[self.line_count, 'Qty']
            if self.df.at[self.line_count, 'Qty'] > 12:
                self.total12 += self.df.at[self.line_count, 'Qty'] - 12
            self.line_count += 1

    def check_60(self):
        """ Checking for over 60 violations, and checking against double-counting """
        i = 1  # More initializations
        self.total60 = self.total - 60
        self.total602 = self.total60
        while self.total602 > 0:
            if self.total602 + 12 <= self.df.at[self.line_count - i, 'Qty']:
                self.total60 -= self.total602
                # self.total602 == -1
            else:
                if float(self.df.at[self.line_count - i, 'Qty']) > 12.00:
                    self.total60 -= float(self.df.at[self.line_count - i, 'Qty'])
                    self.total60 += 12
                self.total602 -= self.df.at[self.line_count - i, 'Qty']
            i += 1

    def calculate(self):
        """ Simple Grievance time calculation """
        if self.total12 > 0 and self.total60 > 0:
            self.trutotal = self.total12 + self.total60
        else:
            if self.total12 > 0:
                self.trutotal = self.total12
            elif self.total60 > 0:
                self.trutotal = self.total60
            else:
                self.trutotal = 0

    def output(self):
        """Output if grievable"""
        format_total = "{:.2f}".format(self.total)  # truncate to 2 decimals
        format_trutotal = "{:.2f}".format(self.trutotal)
        if self.trutotal > .24:
            self.sheet.cell(row=self.xlrow, column=1).value = str(self.df.at[self.line_count, 'Employee_ID'])
            self.sheet.cell(row=self.xlrow, column=2).value = str(self.df.at[self.line_count, 'Last Name'])
            self.sheet.cell(row=self.xlrow, column=3).value = str(format_total)
            self.sheet.cell(row=self.xlrow, column=4).value = str(format_trutotal)
            self.xlrow += 1

    def save(self):
        """ # I like to work hard.  For the post.  I'm postal. """
        self.book.save(self.fileout)
        quit()


if __name__ == "__main__":
    MaxHour().run()
