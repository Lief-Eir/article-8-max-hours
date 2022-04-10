"""
Notes!
I would like to add a way to automatically download the dependencies, as explaining this is awful
I also would like to make it easier for user to load in the file to be parsed
This may include path cleaning.  Logic took like a few hours pathing and un-dummying is going to take forever
"""

# dependencies
from pandas import read_csv
from openpyxl import Workbook

# ask for input.  It is on the user to know.  Make a README
filein = str(input('enter input file name: '))
fileout = str(input('enter desired output file name: '))

# Input/Output Cleaning
if not fileout.endswith('.xlsx'):
    fileout += '.xlsx'
if not filein.endswith('.csv'):
    filein += '.csv'
# Initialize Pandas DataTable, skipping the header
df = read_csv(filein, skiprows=1)

# Initialize output with headers
book = Workbook()
sheet = book.active
sheet.cell(row=1, column=1).value = 'EID'
sheet.cell(row=1, column=2).value = 'NAME'
sheet.cell(row=1, column=3).value = 'TOTAL HOURS'
sheet.cell(row=1, column=4).value = 'GRIEVED HOURS'

# Initialize variables. Line_count and rowmax are for reading, xlrow is for writing
line_count = -1
rowmax = -1
xlrow = 2

# Find the number of rows to iterate over
for row in df.iterrows():
    rowmax += 1

# All the logic
while line_count < rowmax - 4:
    line_count += 1
    total = 0
    total12 = 0

    # Checking craft
    while df.at[line_count, 'D/A'] != 134:
        line_count += 1

    # We are verifying that the row is a full hour row for the same employee, then checking 12 hour violations
    while df.at[line_count, 'Employee_ID'] == df.at[line_count + 1, 'Employee_ID'] and \
            df.at[line_count, 'Hours'] == 'WK':
        total += df.at[line_count, 'Qty']
        if df.at[line_count, 'Qty'] > 12:
            total12 += df.at[line_count, 'Qty'] - 12
        line_count += 1

    # More initializations
    i = 1
    total60 = total - 60
    total602 = total60

    # Checking for over 60 violations, and checking against double-counting
    while total602 > 0:
        if total602 + 12 <= df.at[line_count - i, 'Qty']:
            total60 -= total602
            # total602 == -1
        else:
            if float(df.at[line_count - i, 'Qty']) > 12.00:
                total60 -= float(df.at[line_count - i, 'Qty'])
                total60 += 12
            total602 -= df.at[line_count - i, 'Qty']
        i += 1

    # Simple Grievance time calculation
    if total12 > 0 and total60 > 0:
        trutotal = total12 + total60
    else:
        if total12 > 0:
            trutotal = total12
        elif total60 > 0:
            trutotal = total60
        else:
            trutotal = 0

    # truncate to 2 decimals
    format_total = "{:.2f}".format(total)
    format_trutotal = "{:.2f}".format(trutotal)

    # Output if grievable
    if trutotal > .24:
        sheet.cell(row=xlrow, column=1).value = str(df.at[line_count, 'Employee_ID'])
        sheet.cell(row=xlrow, column=2).value = str(df.at[line_count, 'Last Name'])
        sheet.cell(row=xlrow, column=3).value = str(format_total)
        sheet.cell(row=xlrow, column=4).value = str(format_trutotal)
        xlrow += 1

# I like to work hard.  For the post.  I'm postal.
book.save(fileout)
quit()
